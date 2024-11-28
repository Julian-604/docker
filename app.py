from flask import Flask, render_template, request, redirect, url_for, jsonify
from openpyxl import load_workbook
from datetime import datetime
import os

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your_secret_key'  # for CSRF protection

# Path to your Excel sheet
EXCEL_FILE_PATH = 'database.xlsx'

# Load the Excel sheet
def load_excel():
    if os.path.exists(EXCEL_FILE_PATH):
        return load_workbook(EXCEL_FILE_PATH)
    else:
        wb = load_workbook()
        wb.create_sheet('Data')
        wb.save(EXCEL_FILE_PATH)
        return wb

# Fetch records from Excel sheet
def get_all_records():
    wb = load_excel()
    sheet = wb['Data']
    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        records.append(row)
    return records

# Add a record to the Excel sheet
def add_record(data):
    wb = load_excel()
    sheet = wb['Data']
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append(data + (timestamp,))
    wb.save(EXCEL_FILE_PATH)

# Search for records in Excel sheet
def search_records(query):
    wb = load_excel()
    sheet = wb['Data']
    results = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        if any(query.lower() in str(cell).lower() for cell in row):
            results.append(row)
    return results

# CRUD: Update record in the Excel sheet
def update_record(row_index, data):
    wb = load_excel()
    sheet = wb['Data']
    for col_index, value in enumerate(data, start=1):
        sheet.cell(row=row_index, column=col_index, value=value)
    wb.save(EXCEL_FILE_PATH)

# Route for home page
@app.route('/')
def index():
    return render_template('index.html')

# Route for search functionality
@app.route('/search', methods=['GET'])
def search():
    query = request.args.get('query', '')
    if query:
        records = search_records(query)
        return render_template('index.html', records=records, search_query=query)
    return render_template('index.html', records=[])

# Route to display and edit records
@app.route('/update/<int:row_id>', methods=['GET', 'POST'])
def update(row_id):
    if request.method == 'POST':
        data = [request.form['pis'], request.form['name'], request.form['ip'], request.form['mac'], request.form['switch_ip'], request.form['switch_port'], request.form['building'], request.form['room'], request.form['ext']]
        update_record(row_id, data)
        return redirect(url_for('index'))

    # Fetch record for editing
    wb = load_excel()
    sheet = wb['Data']
    row_data = sheet[row_id]
    return render_template('update.html', record=row_data)

# Route to create a new record
@app.route('/create', methods=['GET', 'POST'])
def create():
    if request.method == 'POST':
        data = [request.form['pis'], request.form['name'], request.form['ip'], request.form['mac'], request.form['switch_ip'], request.form['switch_port'], request.form['building'], request.form['room'], request.form['ext']]
        add_record(data)
        return redirect(url_for('index'))

    return render_template('create.html')

# Route to show recent records
@app.route('/recents')
def recents():
    records = get_all_records()[-5:]  # Last 5 records
    return render_template('recents.html', records=records)

# Route to download all records as Excel
@app.route('/download')
def download():
    return send_file(EXCEL_FILE_PATH, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
